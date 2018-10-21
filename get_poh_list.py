# -*- coding: utf-8 -*-

import os
import time

import openpyxl
import requests
from bs4 import BeautifulSoup

PATH = os.getenv("HOMEDRIVE", "None") + os.getenv("HOMEPATH", "None") + "/Dropbox/POG/"
WBPATH = (PATH + "POG_HorseList.xlsx").replace("\\", "/")


def get_poh_list():
    wb = openpyxl.load_workbook(WBPATH)
    wshl = wb["POHorseList"]

    xlrow = 2
    while wshl.cell(row=xlrow, column=1).value:
        horse_name = wshl.cell(row=xlrow, column=2).value
        horse_name_origin = wshl.cell(row=xlrow, column=3).value

        is_horse_name_determined = False

        if len(horse_name) < 6:
            is_horse_name_determined = True
        elif horse_name[-5] != "の":
            is_horse_name_determined = True

        if is_horse_name_determined and horse_name_origin:
            xlrow += 1
            continue

        horse_url_sp = wshl.cell(row=xlrow, column=5).value
        if horse_url_sp is None:
            xlrow += 1
            continue

        time.sleep(1)
        r = requests.get(horse_url_sp)
        soup = BeautifulSoup(r.content, 'lxml')

        horse_name_new = soup.find("p", class_="Name").string
        horse_name_origin_new = soup.find("th", string="馬名の意味").find_next().string

        wshl.cell(row=xlrow, column=2).value = horse_name_new
        if horse_name_origin_new != "-":
            wshl.cell(row=xlrow, column=3).value = horse_name_origin_new

        xlrow += 1

    wb.save(WBPATH)

    return [[cell.value for cell in row] for row in wshl["A2:F" + str(xlrow - 1)]]


if __name__ == "__main__":
    get_poh_list()
