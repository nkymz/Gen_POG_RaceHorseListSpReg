# -*- coding: utf-8 -*-

import time
from logging import getLogger, StreamHandler, DEBUG

import openpyxl
import requests
from bs4 import BeautifulSoup

logger = getLogger(__name__)
handler = StreamHandler()
handler.setLevel(DEBUG)
logger.setLevel(DEBUG)
logger.addHandler(handler)
logger.propagate = False

# logger.debug('hello')

wb = openpyxl.load_workbook('D:/pdojo/POG_HorseList.xlsx')
wshl = wb["POHorseList"]

age = wshl["B1"].value

trow = 1

while wshl.cell(row=trow, column=1).value is not None:
    # print(trow)

    horseNm = wshl.cell(row=trow, column=2).value
    horseNmOrgn = wshl.cell(row=trow, column=3).value

    isHorseNmDtrmnd = False

    if len(horseNm) < 6:
        isHorseNmDtrmnd = True
    elif horseNm[-5] != "の":
        isHorseNmDtrmnd = True

    if isHorseNmDtrmnd and horseNmOrgn is not None:
        trow += 1
        continue

    horseURLsp = wshl.cell(row=trow, column=5).value
    if horseURLsp is None:
        trow += 1
        continue

    time.sleep(1)
    r = requests.get(horseURLsp)
    soup = BeautifulSoup(r.content, 'lxml')

    horseNmNew = soup.find("p", class_="Name").string
    horseNmOrgnNew = soup.find("th", string="馬名の意味").find_next().string

    wshl.cell(row=trow, column=2).value = horseNmNew
    if horseNmOrgnNew != "-":
        wshl.cell(row=trow, column=3).value = horseNmOrgnNew

    trow += 1

POHList = [[cell.value for cell in row] for row in wshl["A1:F" + str(trow - 1)]]

RHList = []

target_url = 'http://race.netkeiba.com/?rf=navi'
r = requests.get(target_url)  # requestsを使って、webから取得
soup = BeautifulSoup(r.text, 'lxml')  # 要素を抽出

DateList = soup.find('div', class_='DateList_Box')

for DateItem in DateList.find_all('a'):

    if DateItem.get('href').split('=')[-1][0] == 'p':
        continue

    target_url = 'http://race.netkeiba.com' + DateItem.get('href')
    time.sleep(1)
    r = requests.get(target_url)  # requestsを使って、webから取得
    soup = BeautifulSoup(r.content, 'lxml')  # 要素を抽出

    prev_target_url = None

    for race_url in soup.find_all('a'):

        # logger.info(str(race_url.get("href").split("=")[1])[0:4])

        if len(race_url.get("href").split("=")[2].split("&")) > 1:
            raceID = race_url.get("href").split("=")[2].split("&")[0]
        else:
            raceID = race_url.get("href").split("=")[2]

        if raceID[0] != "n":
            continue

        if str(race_url.get("href").split("=")[1])[0:4] != "race":
            continue

        if race_url.get("title")[0:3] == "３歳上" or (age == 2 and race_url.get("title")[0:2] == "３歳"):
            continue

        target_url = 'http://race.netkeiba.com/?pid=race_old&id=' + raceID

        if target_url == prev_target_url:
            continue

        prev_target_url = target_url
        time.sleep(1)
        r = requests.get(target_url)  # requestsを使って、webから取得
        soup = BeautifulSoup(r.content, 'lxml')  # 要素を抽出

        dt_list = soup.find_all('dt', limit=2)

        RaceNo = ("0" + dt_list[1].string.strip())[-3:]

        h1_list = soup.find_all('h1')
        RaceName = h1_list[1].contents[0].strip()
        # print(len(h1_list[1].contents))
        if len(h1_list[1].contents) > 1:
            GradeTemp = str(h1_list[1].contents[1])
            Grade = GradeTemp.split('_')[-2]
        else:
            Grade = ''

        RaceAtrb_list = h1_list[1].find_all_next('p', limit=4)
        course = RaceAtrb_list[0].string.strip()
        RaceCond1 = RaceAtrb_list[2].string
        RaceCond2 = RaceAtrb_list[3].string

        if len(RaceCond1.split()) > 1:
            if RaceCond1.split()[1][0:2] == "障害" or RaceCond1.split()[1][0:3] == "３歳上" or (
                    age == 2 and RaceCond1.split()[1][0:2] == "３歳"):
                continue

        # logger.debug("PASS1")

        title = soup.find('title')
        date = title.string.split()[0]
        track = title.string.split()[1]

        h_list = soup.find_all(class_='bml1')

        for h in h_list:

            HorseName = h.find('td', class_="txt_l horsename").find('div').find('a').string

            isFind = False
            for POHItem in POHList:
                # logger.info(HorseName + " " + POHItem[1])
                if HorseName == POHItem[1]:
                    owner = POHItem[0].strip()
                    isFind = True
                    origin = POHItem[2]
                    if POHItem[5] == "封印":
                        isSeal = True
                    else:
                        isSeal = False

            if not isFind:
                continue

            # logger.debug(h)

            HorseURL = h.find('td', class_="txt_l horsename").find('div').find('a').get('href')
            Weight = h.find('td', class_="txt_l horsename").find_next('td').find_next('td').string
            Jockey = h.find_all('td', class_='txt_l', limit=2)[1].find('a').string
            Odds = h.find('td', class_='txt_r').string
            PopRank = h.find('td', class_='txt_r').find_next('td').string
            SortKey = date + RaceNo + HorseName

            # logger.debug(course)
            # noinspection PyUnboundLocalVariable,PyUnboundLocalVariable,PyUnboundLocalVariable
            RHList.append(
                [SortKey, date, track, RaceNo, RaceName, Grade, course, RaceCond1, RaceCond2,
                 HorseName, Jockey, Odds, PopRank, Weight, target_url, HorseURL, owner, origin, isSeal])

RHList.sort()

f = open("D:/pdojo/SpecialRegistrationList.html", mode="w")

f.write("<p>\n</p>\n")

prevDate = None
prevRaceNo = "00R"

for i in RHList:

    date = i[1]
    track = i[2]
    RaceNo = i[3]
    RaceName = i[4]
    GradeTemp = i[5]
    course = i[6].replace("\xa0", " ")
    RaceCond2 = i[8].replace("\xa0", " ")
    HorseName = i[9]
    Jockey = i[10]
    Odds = i[11]
    PopRank = i[12]
    RaceURL = i[14]
    HorseURL = i[15]
    owner = i[16]
    origin = i[17]
    isSeal = i[18]
    sp = ' '

    if GradeTemp == "g1":
        Grade = "[GI]"
    elif GradeTemp == "g2":
        Grade = "[GII}"
    elif GradeTemp == "g3":
        Grade = "[GIII]"
    elif GradeTemp == "jg1":
        Grade = "[JGI]"
    elif GradeTemp == "jg2":
        Grade = "[JGII]"
    elif GradeTemp == "jg3":
        Grade = "[JGIII]"
    elif GradeTemp == "op":
        Grade = "[OP]"
    else:
        Grade = ""

    s = '<h4>' + date + '</h4>\n'
    if prevDate is None:
        f.write(s)
    elif date != prevDate:
        f.write('</ul></li></ul>' + s)

    s = '<li> <a href="' + RaceURL + '">' + track + RaceNo + sp + RaceName + Grade + '</a><br />\n'
    s2 = course + sp + RaceCond2 + '<br />\n<ul>'
    if prevDate is None or (prevDate is not None and date != prevDate):
        f.write('<ul>' + s)
        f.write(s2)
    elif date + RaceNo != prevDate + prevRaceNo:
        f.write('</ul></li>' + s)
        f.write(s2)

    if isSeal:
        f.write('<li> <a href="' + HorseURL + '"><s>' + HorseName + '</s>' + owner + '</a> <br />\n')
    else:
        f.write('<li> <a href="' + HorseURL + '">' + HorseName + owner + '</a> <br />\n')
    f.write(origin + '<br />\n')
    if Odds is not None:
        f.write(str(Odds) + '倍' + sp + str(PopRank) + '番人気<br />\n')
    if Jockey is not None:
        f.write(Jockey + '騎手<br />\n')
    f.write('</li>\n')

    prevDate = date
    prevRaceNo = RaceNo

f.write('</ul></li></ul><p>終末オーナーLOVEPOP</p>\n')
f.write('<p>※オッズはnetkeibaより取得したものです。</p>')

f.close()

wb.save('POG_HorseList.xlsx')
